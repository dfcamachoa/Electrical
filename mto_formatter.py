import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from design_allowance_manager import get_design_allowance

class MtoFormatter:
    def __init__(self, base_dir):
        """Initialize the MTO formatter with base directory"""
        self.base_dir = base_dir
        self.csv_dir = os.path.join(base_dir, "csv")
        self.output_dir = os.path.join(base_dir, "output")
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # File paths
        self.summed_materials_path = os.path.join(self.csv_dir, "summed_materials.csv")
        self.material_catalog_path = os.path.join(self.csv_dir, "material_catalog.csv")
        self.output_path = os.path.join(self.output_dir, "MTO_Output.xlsx")
        
        # Get design allowance percentage from config
        try:
            self.design_allowance_pct = get_design_allowance(base_dir)
            print(f"MTO Formatter initialized with design allowance: {self.design_allowance_pct}%")
        except Exception as e:
            print(f"Error loading design allowance, using default: {e}")
            self.design_allowance_pct = 10
    
    def load_data(self):
        """Load data from CSV files"""
        # Load summed materials
        self.materials_df = pd.read_csv(self.summed_materials_path)
        
        # Load material catalog for additional info
        self.catalog_df = pd.read_csv(self.material_catalog_path)
        
        # Get unit information from material_by_assembly.csv
        material_by_assembly_path = os.path.join(self.csv_dir, "material_by_assembly.csv")
        if os.path.exists(material_by_assembly_path):
            assembly_materials_df = pd.read_csv(material_by_assembly_path)
            # Create a mapping of TPENG ITEM CODE to UNIT
            self.unit_mapping = assembly_materials_df.drop_duplicates(subset=['TPENG ITEM CODE'])[['TPENG ITEM CODE', 'UNIT']]
            self.unit_mapping = dict(zip(self.unit_mapping['TPENG ITEM CODE'], self.unit_mapping['UNIT']))
        else:
            self.unit_mapping = {}
    
    def prepare_data(self):
        """Prepare and transform data for MTO template"""
        # Ensure TPENG ITEM CODE is treated as string
        self.materials_df['TPENG ITEM CODE'] = self.materials_df['TPENG ITEM CODE'].astype(str)
        self.catalog_df['TPENG ITEM CODE'] = self.catalog_df['TPENG ITEM CODE'].astype(str)
        
        # Get list of unique WBS codes
        self.wbs_codes = sorted(self.materials_df['wbs_code'].unique())
        
        # Get list of unique TPENG ITEM CODEs
        self.item_codes = sorted(self.materials_df['TPENG ITEM CODE'].unique())
        
        # Create a pivot table to reorganize data by item code and WBS
        self.pivot_df = self.materials_df.pivot_table(
            index=['TPENG ITEM CODE', 'BILL OF MATERIAL'],
            columns='wbs_code',
            values='QUANTITY',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        # Add a TOTAL column
        self.pivot_df['TOTAL'] = self.pivot_df[self.wbs_codes].sum(axis=1)
        
        # Add DESIGN ALLOWANCE column with configurable percentage
        self.pivot_df['DESIGN ALLOWANCE'] = (self.pivot_df['TOTAL'] * self.design_allowance_pct / 100).round(2)
        
        # Add GRAND TOTAL column
        self.pivot_df['GRAND TOTAL'] = self.pivot_df['TOTAL'] + self.pivot_df['DESIGN ALLOWANCE']
        
        # Add UNIT column
        self.pivot_df['UNIT'] = self.pivot_df['TPENG ITEM CODE'].map(self.unit_mapping)
        
        # Reorder columns
        column_order = ['TPENG ITEM CODE', 'BILL OF MATERIAL', 'UNIT'] + self.wbs_codes + ['TOTAL', 'DESIGN ALLOWANCE', 'GRAND TOTAL']
        self.pivot_df = self.pivot_df[column_order]
    
    def create_excel(self):
        """Create formatted Excel file based on MTO template"""
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MTO"
        
        # Add headers with additional index column
        headers = ["Item No."] + list(self.pivot_df.columns)
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate positions for project information header
        # We'll place this in the upper-right corner
        total_columns = len(headers)
        project_info_start_col = total_columns - 6  # Start 7 columns from the right
        if project_info_start_col < 8:  # Ensure we have enough space
            project_info_start_col = 8
            
        # Add project information header
        # Row 1: Project info header fields
        project_fields = ["Project No", "Unit Doc.", "Type", "Code", "Serial No", "Rev.", "Page"]
        for i, field in enumerate(project_fields):
            # Labels in row 1
            header_cell = ws.cell(row=1, column=project_info_start_col + i, value=field)
            header_cell.font = Font(bold=True, size=10)
            header_cell.alignment = centered_alignment
            header_cell.border = thin_border
            
            # Empty value cells in row 2
            value_cell = ws.cell(row=2, column=project_info_start_col + i, value="")
            value_cell.border = thin_border
            value_cell.alignment = centered_alignment
        
        # Row 3-5: Additional project info fields
        additional_fields = ["Client Number", "Client Project", "Location", "Unit"]
        for i, field in enumerate(additional_fields):
            # Labels
            header_cell = ws.cell(row=3 + i, column=project_info_start_col, value=field)
            header_cell.font = Font(bold=True, size=10)
            header_cell.alignment = left_alignment
            header_cell.border = thin_border
            
            # Empty value cells spanning across the remaining columns
            ws.merge_cells(start_row=3 + i, start_column=project_info_start_col + 1, 
                          end_row=3 + i, end_column=project_info_start_col + 6)
            value_cell = ws.cell(row=3 + i, column=project_info_start_col + 1, value="")
            value_cell.border = thin_border
        
        # Add title for MTO section (now moved to row 7)
        mto_start_row = 7
        title_cell = ws.cell(row=mto_start_row, column=1, value=f"Material Take-Off (Design Allowance: {self.design_allowance_pct}%)")
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=mto_start_row, start_column=1, end_row=mto_start_row, end_column=len(headers))
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write header row (now at row 8)
        header_row = mto_start_row + 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = thin_border
        
        # Write data rows (starting at row 9)
        data_start_row = mto_start_row + 2
        for row_idx, row in enumerate(self.pivot_df.itertuples(index=False), data_start_row):
            # Add index number in first column
            index_cell = ws.cell(row=row_idx, column=1, value=row_idx - data_start_row + 1)  # Index starts at 1
            index_cell.border = thin_border
            index_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data in remaining columns
            for col_idx, value in enumerate(row, 2):  # Start at column 2 since column 1 is the index
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Right-align numeric columns
                if col_idx > 4:  # After Item No., TPENG ITEM CODE, BILL OF MATERIAL, UNIT
                    cell.alignment = Alignment(horizontal='right')
                    # Format as integer if it's a whole number
                    if isinstance(value, (int, float)) and value == int(value):
                        cell.value = int(value)
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Highlight Design Allowance and Grand Total columns
        design_col = headers.index('DESIGN ALLOWANCE') + 1
        grand_total_col = headers.index('GRAND TOTAL') + 1
        
        highlight_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row in range(data_start_row, len(self.pivot_df) + data_start_row):
            ws.cell(row=row, column=design_col).fill = highlight_fill
            ws.cell(row=row, column=grand_total_col).fill = highlight_fill
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Find the maximum length in the column
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Adjust column width (with some padding)
            adjusted_width = max_length + 2
            
            # Set minimum and maximum widths
            if col_idx == 1:  # Item No. column
                adjusted_width = 8  # Fixed width for index column
            elif col_idx == 3:  # BILL OF MATERIAL column
                adjusted_width = max(adjusted_width, 40)  # Minimum width for description
            else:
                adjusted_width = min(max(adjusted_width, 10), 30)  # Between 10 and 30
                
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = f'A{data_start_row}'  # Freeze after title and header rows
        
        # Save the workbook
        wb.save(self.output_path)
        
        return self.output_path
    
    def run(self):
        """Run the MTO formatting process"""
        try:
            self.load_data()
            self.prepare_data()
            output_path = self.create_excel()
            return True, f"MTO formatted successfully with {self.design_allowance_pct}% design allowance. Output saved to: {output_path}"
        except Exception as e:
            return False, f"MTO formatting failed: {str(e)}"


def format_mto(base_dir):
    """Run MTO formatting from base directory"""
    formatter = MtoFormatter(base_dir)
    return formatter.run()


# Example usage:
if __name__ == "__main__":
    # Base directory is the current directory
    base_dir = os.getcwd()
    success, message = format_mto(base_dir)
    print(message)